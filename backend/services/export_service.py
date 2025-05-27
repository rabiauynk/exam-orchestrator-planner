import os
import tempfile
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from database import db
from models import ExamSchedule, Exam, Department, Room

class ExportService:
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def export_all_departments_excel(self, start_date=None, end_date=None):
        """Export all departments' exam schedules to a single Excel file with multiple sheets"""
        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Get all departments
            departments = Department.query.all()
            
            for department in departments:
                # Get department's exam schedules
                data = self._get_department_schedule_data(department.id, start_date, end_date)
                
                if data:
                    # Create sheet for department
                    ws = wb.create_sheet(title=department.code)
                    self._format_department_sheet(ws, department, data)
            
            # If no data found, create an empty sheet
            if not wb.worksheets:
                ws = wb.create_sheet(title="No Data")
                ws['A1'] = "No exam schedules found for the specified date range."
            
            # Save file
            filename = f"exam_schedule_all_departments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.temp_dir, filename)
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            print(f"Error exporting all departments Excel: {str(e)}")
            return None
    
    def export_department_excel(self, department_id, start_date=None, end_date=None):
        """Export specific department's exam schedule to Excel"""
        try:
            # Get department
            department = Department.query.get(department_id)
            if not department:
                return None
            
            # Get department's exam schedules
            data = self._get_department_schedule_data(department_id, start_date, end_date)
            
            if not data:
                # Create empty file
                wb = Workbook()
                ws = wb.active
                ws.title = department.code
                ws['A1'] = "No exam schedules found for the specified date range."
            else:
                # Create workbook with data
                wb = Workbook()
                ws = wb.active
                ws.title = department.code
                self._format_department_sheet(ws, department, data)
            
            # Save file
            filename = f"exam_schedule_{department.code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.temp_dir, filename)
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            print(f"Error exporting department Excel: {str(e)}")
            return None
    
    def _get_department_schedule_data(self, department_id, start_date=None, end_date=None):
        """Get exam schedule data for a department"""
        try:
            # Build query
            query = db.session.query(
                ExamSchedule.scheduled_date,
                ExamSchedule.start_time,
                ExamSchedule.end_time,
                Exam.course_name,
                Exam.class_name,
                Exam.instructor,
                Exam.student_count,
                Exam.duration,
                Exam.needs_computer,
                Room.name.label('room_name')
            ).join(Exam).join(Room).filter(
                Exam.department_id == department_id
            )
            
            # Apply date filters
            if start_date:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(ExamSchedule.scheduled_date >= start_date_obj)
            
            if end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(ExamSchedule.scheduled_date <= end_date_obj)
            
            # Order by date and time
            schedules = query.order_by(
                ExamSchedule.scheduled_date.asc(),
                ExamSchedule.start_time.asc()
            ).all()
            
            # Convert to list of dictionaries
            data = []
            for schedule in schedules:
                data.append({
                    'Tarih': schedule.scheduled_date.strftime('%d/%m/%Y'),
                    'Başlangıç Saati': schedule.start_time.strftime('%H:%M'),
                    'Bitiş Saati': schedule.end_time.strftime('%H:%M'),
                    'Ders Adı': schedule.course_name,
                    'Sınıf': schedule.class_name,
                    'Öğretim Görevlisi': schedule.instructor,
                    'Öğrenci Sayısı': schedule.student_count,
                    'Süre (dk)': schedule.duration,
                    'Bilgisayar Gerekli': 'Evet' if schedule.needs_computer else 'Hayır',
                    'Sınıf/Lab': schedule.room_name
                })
            
            return data
            
        except Exception as e:
            print(f"Error getting department schedule data: {str(e)}")
            return []
    
    def _format_department_sheet(self, ws, department, data):
        """Format Excel sheet for department data"""
        try:
            # Set title
            ws['A1'] = f"{department.name} - Sınav Programı"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:J1')
            
            # Add generation date
            ws['A2'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(size=10, italic=True)
            ws.merge_cells('A2:J2')
            
            # Add empty row
            current_row = 4
            
            if data:
                # Convert to DataFrame for easier handling
                df = pd.DataFrame(data)
                
                # Add headers
                headers = list(df.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                current_row += 1
                
                # Add data rows
                for row_data in data:
                    for col_num, value in enumerate(row_data.values(), 1):
                        cell = ws.cell(row=current_row, column=col_num, value=value)
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    current_row += 1
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            else:
                ws['A4'] = "Bu bölüm için belirtilen tarih aralığında sınav bulunmamaktadır."
                ws['A4'].font = Font(italic=True)
                ws.merge_cells('A4:J4')
            
        except Exception as e:
            print(f"Error formatting department sheet: {str(e)}")
    
    def cleanup_temp_files(self, max_age_hours=24):
        """Clean up old temporary export files"""
        try:
            current_time = datetime.now()
            for filename in os.listdir(self.temp_dir):
                if filename.startswith('exam_schedule_') and filename.endswith('.xlsx'):
                    file_path = os.path.join(self.temp_dir, filename)
                    file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                    age_hours = (current_time - file_time).total_seconds() / 3600
                    
                    if age_hours > max_age_hours:
                        try:
                            os.remove(file_path)
                            print(f"Cleaned up old export file: {filename}")
                        except OSError:
                            pass
        except Exception as e:
            print(f"Error cleaning up temp files: {str(e)}")
